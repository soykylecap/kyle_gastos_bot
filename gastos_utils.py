from telegram import Update
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes
import openpyxl
import xlwings as xw
import logging
import os
from datetime import datetime
from dotenv import load_dotenv

# ─── CONFIGURACIÓN ───────────────────────────────────────────────
load_dotenv()

TOKEN = os.getenv("TOKEN")
HOJA = os.getenv("HOJA")
EXCEL_PATH = os.getenv("EXCEL_PATH")
EXCEL_PAGOS = os.getenv("EXCEL_PAGOS")
CANTIDAD_ULTIMOS_MOVIMIENTOS = int(os.getenv("CANTIDAD_ULTIMOS_MOVIMIENTOS"))
RUBROS = os.getenv("RUBROS").split(", ")

# ─────────────────────────────────────────────────────────────────

logging.basicConfig(level=logging.INFO)

def consulta_saldo():
    app = xw.App(visible=False)
    wb = xw.Book(EXCEL_PATH)
    ws = wb.sheets[HOJA]
    saldo_pesos = int(ws.range("M1").value)
    saldo_dolares = int(ws.range("M2").value)
    wb.save()
    wb.close()
    app.quit()
    return saldo_pesos, saldo_dolares

    # A continuacion codigo para consultar en servidor. No muestra ultimos cambios
    # Seria mejor sumar todos los ingresos y todos los egregsos y con eso calcular saldo
    #
    # wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    # ws = wb[HOJA]
    # saldo_pesos = int(ws["M1"].value)
    # saldo_dolares = int(ws["M2"].value)
    # wb.close()
    # return saldo_pesos, saldo_dolares


def consulta_pagos():
    app = xw.App(visible=False)
    wb = xw.Book(EXCEL_PAGOS)
    ws = wb.sheets["Hoja1"]
    debe = int(ws.range("D36").value)
    haber = int(ws.range("E36").value)
    wb.save()
    wb.close()
    app.quit()
    return debe, haber


def parsear_mensaje(texto):
    palabras = texto.split()
    monto = None
    for palabra in reversed(palabras):
        monto_original = palabra
        palabra_limpia = palabra.replace(".", "").replace(",", "")
        if palabra_limpia.isdigit():
            monto = float(palabra_limpia)
            break

    if not monto:
        return None, None
    detalle = texto
    detalle = detalle.replace(str(monto_original), "").strip()
    detalle = " ".join(detalle.split())

    return detalle, monto


def ultimo_renglon(columna):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[HOJA]
    valor_max = 0
    primer_renglon_datos = 2
    ultimo_renglon_ocupado = primer_renglon_datos - 1
    for fila in ws.iter_rows(min_row=primer_renglon_datos):
        celda = fila[columna]

        if celda.value is not None and isinstance(celda.value, (int, float)) and celda.value > valor_max:
            valor_max = celda.value
            ultimo_renglon_ocupado = celda.row
    return ultimo_renglon_ocupado, valor_max


def ultimos_movimientos(cantidad):

    ultimo_renglon_ocupado, _ = ultimo_renglon(1)
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[HOJA]
    campo = list()

    for r in reversed(range(cantidad)):
        renglon = ultimo_renglon_ocupado - r
        fecha = ws.cell(row=renglon, column=3).value
        detalle = ws.cell(row=renglon, column=5).value
        rubro = ws.cell(row=renglon, column=6).value
        ingreso = ws.cell(row=renglon, column=8).value
        gasto = ws.cell(row=renglon, column=9).value
        campo.append((fecha, detalle, rubro, ingreso, gasto))
    return campo



def registrar_en_excel(fecha, cantidad, detalle, rubro, cotiza_dolar, ingreso, gasto, autor):
    ultimo_renglon_ocupado, valor_max = ultimo_renglon(1)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[HOJA]
    ws.cell(row=ultimo_renglon_ocupado + 1, column=2, value=valor_max + 1)   # PK
    ws.cell(row=ultimo_renglon_ocupado + 1, column=3, value=fecha)
    ws.cell(row=ultimo_renglon_ocupado + 1, column=4, value=cantidad)  # Cantidad
    ws.cell(row=ultimo_renglon_ocupado + 1, column=5, value=detalle)
    ws.cell(row=ultimo_renglon_ocupado + 1, column=6, value=rubro)
    ws.cell(row=ultimo_renglon_ocupado + 1, column=7, value=cotiza_dolar)   # Dolar
    ws.cell(row=ultimo_renglon_ocupado + 1, column=8, value=ingreso)   # Ingreso
    ws.cell(row=ultimo_renglon_ocupado + 1, column=9, value=gasto)
    ws.cell(row=ultimo_renglon_ocupado + 1, column=12, value=autor)  

    if rubro == "Dolares":
        ultimo_renglon_ocupado, valor_max = ultimo_renglon(13)

        ws.cell(row=ultimo_renglon_ocupado + 1, column=14, value=valor_max + 1)   # PK
        ws.cell(row=ultimo_renglon_ocupado + 1, column=15, value=fecha)
        ws.cell(row=ultimo_renglon_ocupado + 1, column=16, value=detalle)  # detalle
        cantidad = int(cantidad) *-1
        if autor == "Andrés":
            ws.cell(row=ultimo_renglon_ocupado + 1, column=20, value=cantidad)
        else:
            ws.cell(row=ultimo_renglon_ocupado + 1, column=18, value=cantidad)
    wb.save(EXCEL_PATH)


async def cancelar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await context.bot.send_message(chat_id=update.message.chat_id, text="❌ Operación cancelada.")


async def saldo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    autor = update.message.from_user.first_name
    await update.message.reply_text(f"¡Hola {autor}! Bancame que hago unas cuentas y te digo 🤓")
    saldo_pesos, saldo_dolares = consulta_saldo()
    await update.message.reply_text(f"💰 Saldos:\n$ {saldo_pesos:,.0f}\nu$s {saldo_dolares:,.0f}")


async def pagos(update: Update, context: ContextTypes.DEFAULT_TYPE):
    autor = update.message.from_user.first_name
    await update.message.reply_text(f"¡Hola {autor}! Ahora te digo cuanto queda pendiente. Hay que cuidar el 🐷")
    debe, haber = consulta_pagos()
    await update.message.reply_text(f"Presupuesto: u$s {debe:,.0f}\nYa pagaron: u$s {haber:,.0f}\nPendiente: u$s {debe-haber:,.0f}")


async def dolar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    autor = update.message.from_user.first_name
    context.user_data["autor"] = autor
    context.user_data["rubro"] = "Dolares"
    context.user_data["esperando_cantidad"] = True
    context.user_data["esperando_rubro"] = True
    await update.message.reply_text(f"¿Cuántos dólares vendiste?")


async def ultimos(update: Update, context: ContextTypes.DEFAULT_TYPE):
    campo = ultimos_movimientos(CANTIDAD_ULTIMOS_MOVIMIENTOS)
    
    mensaje = "```\n"
    mensaje += f"{'Fecha':<5} {'Detalle':<9} {'Rubro':<4} {'Ingre':<7} {'Gasto':>7}\n"
    for r in range(CANTIDAD_ULTIMOS_MOVIMIENTOS):
        fecha = f"{campo[r][0]:%d/%m}"
        detalle = campo[r][1][:9] or ""
        rubro = campo[r][2][:4] or ""
        ingreso = campo[r][3] or 0
        gasto = campo[r][4] or 0
        mensaje += f"{fecha:<5} {detalle:<9} {rubro:<4} {ingreso:>7} {gasto:>7}\n"
    mensaje += "```"

    await context.bot.send_message(chat_id=update.message.chat_id, text=mensaje, parse_mode="Markdown")


async def gasto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    autor = update.message.from_user.first_name
    await update.message.reply_text(f"¡Hola {autor}! Ingresa el detalle y el monto (ej: Tornillos Rothoblaas 130.000)")


async def manejar_mensaje(update: Update, context: ContextTypes.DEFAULT_TYPE):
    logging.info(f">>> manejar_mensaje llamado con: {update.message.text}")

    texto = update.message.text
    autor = update.message.from_user.first_name
    chat_id = update.message.chat.id

    print (type(chat_id))

    if "esperando_rubro" not in context.user_data:
        detalle, monto = parsear_mensaje(texto)
        logging.info(f">>> detalle: {detalle}, monto: {monto}")
        context.user_data["detalle"] = detalle 
        context.user_data["monto"] = monto
        context.user_data["esperando_rubro"] = True
        context.user_data["autor"] = autor
        filas = []
        columnas = 4
        for i in range(0, len(RUBROS), columnas):
            fila = [InlineKeyboardButton(r, callback_data=r) for r in RUBROS[i:i+columnas]]
            filas.append(fila)
        teclado = InlineKeyboardMarkup(filas)
        await update.message.reply_text("¿A qué rubro pertenece?", reply_markup=teclado)


    elif "esperando_cantidad" in context.user_data:
        cantidad_dolares = update.message.text
        context.user_data["cantidad_dolares"] = cantidad_dolares
        context.user_data.pop("esperando_cantidad")
        context.user_data["esperando_cotizacion"] = True
        await update.message.reply_text(f"¿A que cotizacion?")


    elif "esperando_cotizacion" in context.user_data:
        cotiza_dolar = int(update.message.text)
        cantidad_dolar = context.user_data["cantidad_dolares"]
        monto = int(cotiza_dolar) * int(cantidad_dolar)
        gasto = None
        fecha = datetime.now()
        detalle = f"Vendimos u$s {cantidad_dolar} a $ {cotiza_dolar}."
        rubro = context.user_data["rubro"]
        fecha_format = datetime.now().strftime("%d/%m/%y")
        autor = context.user_data["autor"]
        opciones = [[InlineKeyboardButton(callback_data="Cancelar", text="Cancelar"), InlineKeyboardButton(callback_data="Confirmar", text="Confirmar")]]
        teclado_ok = InlineKeyboardMarkup(opciones)

        await update.message.reply_text(
            f"✅ Venta de dolares registrada:\n"
            f"📅 Fecha: {fecha_format}\n"
            f"📝 Detalle: Vendimos u$s {cantidad_dolar} a $ {cotiza_dolar}\n"
            f"🏷️ Rubro: {rubro}\n"
            f"💰 Ingreso: ${monto:,.0f}\n"
            f"😉 Autor: {autor}\n"
            #, reply_markup=teclado_ok
        )
        
        registrar_en_excel(fecha, cantidad_dolar, detalle, rubro, cotiza_dolar, monto, gasto, autor)
        context.user_data.pop("esperando_rubro")
        context.user_data.pop("esperando_cotizacion")
        context.user_data.pop("cantidad_dolares")
        context.user_data.pop("autor")
        context.user_data.pop("rubro")


async def manejar_boton(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if "esperando_ok" in context.user_data:
        pass
    
    elif "esperando_ok" not in context.user_data:
        query = update.callback_query
        rubro = query.data  # acá llega el callback_data del botón tocado
        await query.answer()  # esto le dice a Telegram que el callback fue recibido
        fecha = datetime.now()
        fecha_format = datetime.now().strftime("%d/%m/%y")
        cantidad = None
        detalle = context.user_data["detalle"]
        cotiza_dolar = None
        ingreso = None
        gasto = context.user_data["monto"]
        autor = context.user_data["autor"]
        
        await query.message.delete()
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=f"✅ Gasto registrado:\n"
                f"📅 Fecha: {fecha_format}\n"
                f"📝 Detalle: {detalle}\n"
                f"🏷️ Rubro: {rubro}\n"
                f"💰 Monto: ${gasto:,.0f}\n"
                f"😉 Autor: {autor}\n"
        )

        registrar_en_excel(fecha, cantidad, detalle, rubro, cotiza_dolar, ingreso, gasto, autor)
        context.user_data.pop("esperando_rubro")
        context.user_data.pop("detalle")
        context.user_data.pop("monto")
        context.user_data.pop("autor")